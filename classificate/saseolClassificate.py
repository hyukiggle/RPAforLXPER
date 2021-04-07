# 사설모의고사를 유형별로 분류해서 엑셀에 저장하는 코드

from openpyxl import Workbook
import os

choice = input("일괄처리면 1, 단일파일이면 0 : ")




def qCheck(fName):
    totalLines = []
    totalIndex = []
    sharpIndex = []
    questions = []
    contents = []
    choices = []
    typeList = []
    qNumList = []
    starWords = []
    threePoints = []
    onec=[]
    twoc=[]
    threec=[]
    fourc=[]
    fivec=[]
    refs = []
    answers = []
    n = 0
    type_elements = ['분위기','심경','변화','요지','주제','제목','주장','빈칸','요약']
    for line in fName.readlines():
        totalLines.append(line.strip())
        totalIndex.append(n)
        if('##' in line):
            sharpIndex.append(n)

        n+=1

    for s in range(len(sharpIndex)-3):

        #ws.cell(row=s+2,column=12).value =
        if (s % 4 == 0):
            question = totalLines[sharpIndex[s]+1]
            qtype =""
            qNum = question[:2]
            answer = question[-1]
            mark = question[-2]

            answers.append(answer)
            if(('분위기' in question) or ('심경' in question) or ('변화'in question)):
                qtype = '심경_분위기_변화'
                question = question[4:-1]
            elif('요지' in question):
                qtype = '요지'
                question = question[4:-1]
            elif('주제' in question):
                qtype = '주제'
                question = question[4:-1]
            elif('제목' in question):
                qtype = '제목'
                question = question[4:-1]
            elif('주장' in question):
                qtype = '주장'
                question = question[4:-1]
            elif(('빈칸' in question) and ('요약' not in question)):
                if('(B)에 들어갈' in question):
                    if mark == 'T':
                        qtype = '연결사'
                else:
                    if mark == '@':
                        qtype = '빈칸(단어)'
                    else:
                        qtype = '빈칸(구,절)'
                question = '다음 글의 빈칸에 들어갈 말로 가장 적절한 것은?'
                '''
                if('@' in qustion):
                   qtype = '빈칸(단어)'
                elif('(B)에 들어갈' in question):
                    qtype = '빈칸(AB)'
                '''
            elif('요약' in question):
                if mark == '#':
                    qtype = '요약(구,절)'
                else:
                    qtype = '요약'
                question = '다음 글의 내용을 한 문장으로 요약하고자 한다. 빈칸 (A)와 (B)에 들어갈 말로 가장 적절한 것은?'
            else:
                qtype = 'Unknown'
            qNumList.append(qNum)
            questions.append(question)
            #print(qtype, question)
            typeList.append(qtype)
        elif (s % 4 == 1):
            content = ""
            for j in range(sharpIndex[s]+1, sharpIndex[s+1]):
                content += totalLines[j] + " "
            '''
            for c in range(len(content)):
                if(content[c:c+1] == '*'):
                    #ws.cell(row=s + 1, column=11).value = content[c:]
                    starWords.append(content[c:])
                    content = content[:c]
            '''

            contents.append(content)

        elif(s%4 == 2):
            for j in range(sharpIndex[s]+1, sharpIndex[s + 1]):
                choices.append(totalLines[j].strip())
                if('\n' in choices):
                    choices.remove('\n')
                elif(''in choices):
                    choices.remove('')
    '''
    for ct in range(len(contents)):
        if(('[3' or '〔3') in contents[ct]):
            threePoints.append('3')
        else:
            threePoints.append('2')
        if('*' in contents[ct] ):
            for st in range(len(contents[ct])):
                if(contents[ct][st:st+1] == '*'):
                    starWords.append(contents[ct][st:].strip())
        else:
            starWords.append("")
    '''
    for q in range(len(contents)):
        if ('3점' or '3 점') in contents[q]:
            threePoints.append(3)
        else:
            threePoints.append(2)
        if '*' in contents[q]:
           for n in range(len(contents[q])):
                if (contents[q][n:n+3] == '. *' or contents[q][n:n+3] == '] *' or contents[q][n:n+3] == '! *'):
                   starWords.append(contents[q][n+2:].strip())
        else:
            starWords.append("")

    print(len(starWords),len(questions), len(answers))
    print(typeList)

    for c in range(len(choices)):
        if (c % 5 == 0):
            onec.append(choices[c][1:])
        elif (c % 5 == 1):
            twoc.append(choices[c][1:])
        elif (c % 5 == 2):
            threec.append(choices[c][1:])
        elif (c % 5 == 3):
            fourc.append(choices[c][1:])
        elif (c % 5 == 4):
            fivec.append(choices[c][1:])
    #print(choices)
    #print("문제 수 " , len(contents))


    # 요약문장, 보기 넣기
    sumChoice = []
    summary = ""
    for sum in range(sharpIndex[-3]+1,sharpIndex[-2]):
        summary += totalLines[sum].strip()
    for k in range(sharpIndex[-2]+1,sharpIndex[-1]):
        sumChoice.append(totalLines[k].strip())
        for c in range(len(sumChoice)):
            if ('\n' in sumChoice):
                sumChoice.remove('\n')
            elif ('' in sumChoice):
                sumChoice.remove('')

    wb = Workbook()
    ws = wb.active
    ws.append(['번호', '유형', '문제', '지문', '정답', '보기1', '보기2',
               '보기3', '보기4', '보기5', '별단어', '출처', '요약문', '', '점수', '비고'])

    def cellInput(list, num):
        for i in range(len(list)):
            ws.cell(row=i + 2, column=num).value = list[i]

    ws.cell(row=len(qNumList)+1,column=6).value = sumChoice[0][1:]
    ws.cell(row=len(qNumList)+1,column=7).value = sumChoice[1][1:]
    ws.cell(row=len(qNumList)+1,column=8).value = sumChoice[2][1:]
    ws.cell(row=len(qNumList)+1,column=9).value = sumChoice[3][1:]
    ws.cell(row=len(qNumList)+1,column=10).value = sumChoice[4][1:]
    ws.cell(row=len(qNumList)+1, column=13).value = summary



    cellInput(qNumList,1)
    cellInput(typeList, 2)
    cellInput(questions,3)
    cellInput(contents,4)
    cellInput(answers,5)
    cellInput(onec, 6)
    cellInput(twoc, 7)
    cellInput(threec, 8)
    cellInput(fourc, 9)
    cellInput(fivec, 10)
    cellInput(starWords,11)
    cellInput(threePoints,15)

    number = len(questions)
    refs = []
    for i in range(number):
        refs.append(files[:-4])
    cellInput(refs, 12)
    wb.save(filename="C:/Users/LXPER MINI001/Desktop/잡업/temp/고1,2/{}.xlsx".format(files[:-4]))
    return 0

if(choice == '1'):
    os.chdir("C:/Users/LXPER MINI001/Desktop/잡업/temp/고1,2")
    filelist = os.listdir()
    filelist.sort()
    for files in filelist:
        if files.endswith(".txt"):
            filename = open('C:/Users/LXPER MINI001/Desktop/잡업/temp/고1,2/{}'.format(files),'r', encoding='utf-8')
            print(files,"작업 시작")
            qCheck(filename)

            print(files,"작업 끝")



else:
    file = open("//lxper-share/LXPER공유폴더/이승혁/고3 2012년 사설_엑셀완료/고3-2012년 3월 중앙.txt",'r', encoding='utf-8')
    qCheck(file)
    for i in range(16):
        ws.cell(row=i+2,column=12).value = '2012년 4월 고3 종로'
    wb.save(filename="C:/Users/LXPER MINI001/Desktop/잡업/temp/고3-2012년 3월 중앙.xlsx")