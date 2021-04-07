from openpyxl import Workbook
import os
choice = input("일괄처리면 1, 단일파일이면 0: ") #str타입 반환


def classificate(filename):
    typeList = ['분위기','심경','심정','요지','주제','제목','주장','빈칸','요약']
    totalLines = []
    totalIndex = []
    qIndex = []
    qNum = []
    choiceIndex = []
    questions = []
    answers = []
    qTypes = []
    contents = []
    ones = []
    twos = []
    threes = []
    fours = []
    fives = []
    points = []
    asteroids = []
    n = 0


    for line in filename.readlines():
        totalLines.append(line.strip())
        totalIndex.append(n)
        n += 1

    for t in range(len(totalLines)):
        for elem in typeList:
            if elem in totalLines[t] and '다음' in totalLines[t] :
                line = totalLines[t].strip()
                questions.append(line[4:-1].strip())
                qNum.append(line[:2])
                qIndex.append(t)
                answers.append(line[-1])

        if('①' in totalLines[t]):
            ones.append(totalLines[t][2:].strip())
            choiceIndex.append(t)
        elif('②'in totalLines[t]): twos.append(totalLines[t][2:].strip())
        elif('③'in totalLines[t]): threes.append(totalLines[t][2:].strip())
        elif('④'in totalLines[t]): fours.append(totalLines[t][2:].strip())
        elif('⑤'in totalLines[t]): fives.append(totalLines[t][2:].strip())
    for question in questions:
        mark = question[-1]
        if ('분위기' in question): qtype = '심경_분위기_변화'
        elif ('심경' in question): qtype = '심경_분위기_변화'
        elif('변화' in question): qtype = '심경_분위기_변화'
        elif ('요지' in question): qtype = '요지'
        elif ('주제' in question): qtype = '주제'
        elif ('제목' in question): qtype = '제목'
        elif ('주장' in question): qtype = '주장'
        elif (('빈칸' in question) and ('요약' not in question)):
            if mark == 'T':
                qtype = '연결사'
            elif mark == '@':
                qtype = '빈칸(단어)'
            else:
                qtype = '빈칸(구,절)'
        elif ('요약' in question):
            qtype = '요약'
            if ('#' in question):
                qtype = '요약(구,절)'
        else: qtype ='unknown'
        qTypes.append(qtype)
    new_qIndex = []
    for index in qIndex:
        if index not in new_qIndex:
            new_qIndex.append(index)


    questions.pop()
    qNum.pop()
    answers.pop()
    qIndex.pop()
    qTypes.pop()
    print(len(qIndex), qIndex)
    print(len(new_qIndex), new_qIndex)
    print(len(choiceIndex), choiceIndex)
    print(len(qTypes), qTypes)


    for i in range(len(new_qIndex)):
        article = ""
        for j in range(new_qIndex[i]+1, choiceIndex[i]):
           article += totalLines[j].strip() + " "

        contents.append(article)

    for q in range(len(contents)):
        if ('3점' or '3 점') in contents[q]:
            points.append(3)
        else:
            points.append(2)
        if '*' in contents[q]:
           for n in range(len(contents[q])):
                if (contents[q][n:n+3] == '. *' or contents[q][n:n+3] == '] *' or contents[q][n:n+3] == '! *'):
                   asteroids.append(contents[q][n+2:].strip())

        else:
            asteroids.append("")
    summary = ""
    for s in range(len(contents[-1])):

        if(contents[-1][s:s+1] == '↓'):
            summary = contents[-1][s:]
            contents[-1] = contents[-1][:s]

    wb = Workbook()
    ws = wb.active
    ws.append(['번호', '유형', '문제', '지문', '정답', '보기1', '보기2',
               '보기3', '보기4', '보기5', '별단어', '출처', '요약문', '', '점수', '비고'])

    ws.cell(row=len(contents)+1,column=13).value = summary
    refs = []
    for i in range(len(questions)):
        refs.append(files[:-4])


    def cellInput(list, num):
        for i in range(len(list)):
            ws.cell(row=i + 2, column=num).value = list[i]

    cellInput(qNum,1)
    cellInput(qTypes,2)
    cellInput(questions,3)
    cellInput(contents,4)
    cellInput(answers,5)
    cellInput(ones,6)
    cellInput(twos,7)
    cellInput(threes,8)
    cellInput(fours,9)
    cellInput(fives,10)
    cellInput(asteroids,11)
    cellInput(refs, 12)

    cellInput(points,15)

    #print(contents)


    wb.save(filename="//lxper-share/LXPER공유폴더/이승혁/SS txt_238/OCR9/{}.xlsx".format(files[:-4]))
    return 0

if(choice == '1'):
    filelist = os.listdir("//lxper-share/LXPER공유폴더/이승혁/SS txt_238/OCR9")
    filelist.sort()
    for files in filelist:
        if files.endswith(".txt"):
            filename = open('//lxper-share/LXPER공유폴더/이승혁/SS txt_238/OCR9/{}'.format(files),'r', encoding='utf-8')
            print(files, "작업 시작")

            classificate(filename)
            print(files, "작업 끝")


else:
    file = open("//lxper-share/LXPER공유폴더/이승혁/SS txt_238/OCR10/2012년 5월 고1 종로_OCR10.txt",'r',encoding='utf8')
    print("Classificate start")
    classificate(file)
    refs = []
    number = classificate(file)
    for i in range(number):
        refs.append(file[:-4])
    print("classsificate ended")
    wb.save(filename="C:/Users/LXPER MINI001/Desktop/잡업/temp/2012년 5월 고1 종로_OCR10.xlsx")