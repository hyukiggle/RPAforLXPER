from openpyxl import Workbook
from openpyxl import load_workbook
import os

fName = input("저장할 파일 이름을 입력하세요: ")
'''
# pandas
path = r'C:/Users/LXPER MINI001/Desktop/잡업/고2 2014년 사설'
data = pd.read_excel(path)

'''
wb = Workbook()
ws1 = wb.active

ws1.title = '주장'
ws1.append(['번호','유형','문제','지문','정답','보기1','보기2','보기3','보기4','보기5','별단어','출처'])
ws2 = wb.create_sheet('요지')
ws2.append(['번호','유형','문제','지문','정답','보기1','보기2','보기3','보기4','보기5','별단어','출처'])
ws3 = wb.create_sheet('주제')
ws3.append(['번호','유형','문제','지문','정답','보기1','보기2','보기3','보기4','보기5','별단어','출처'])
ws4 = wb.create_sheet('제목')
ws4.append(['번호','유형','문제','지문','정답','보기1','보기2','보기3','보기4','보기5','별단어','출처'])
ws5 = wb.create_sheet('심경_분위기_변화')
ws5.append(['번호','유형','문제','지문','정답','보기1','보기2','보기3','보기4','보기5','별단어','출처'])
ws6 = wb.create_sheet('빈칸(구,절)')
ws6.append(['번호','유형','문제','지문','정답','보기1','보기2','보기3','보기4','보기5','별단어','출처'])
ws7 = wb.create_sheet('빈칸(단어)')
ws7.append(['번호','유형','문제','지문','정답','보기1','보기2','보기3','보기4','보기5','별단어','출처'])
ws8 = wb.create_sheet('연결사')
ws8.append(['번호','유형','문제','지문','정답','보기1','보기2','보기3','보기4','보기5','별단어','출처'])
ws9 = wb.create_sheet('요약(구,절)')
ws9.append(['번호','유형','문제','지문','정답','보기1','보기2','보기3','보기4','보기5','별단어','출처','요약문'])
ws10 = wb.create_sheet('요약(단어)')
ws10.append(['번호','유형','문제','지문','정답','보기1','보기2','보기3','보기4','보기5','별단어','출처','요약문'])


os.chdir("C:/Users/LXPER MINI001/Desktop/잡업/고2 2014년 사설")
filelist = os.listdir()
for files in filelist:
    if files.endswith(".xlsx"):

        load_wb = load_workbook(files, data_only=True)
        load_ws = load_wb['Sheet']


        typeList = []
        for row in load_ws.rows:
            typeList.append(row[1].value)

        for row in load_ws.iter_rows(min_row=2,max_row=10,min_col=1,max_col=13):
            temp = []
            for cell in row:
                temp.append(cell.value)

            print(temp)
            qType = temp[1]
            if(qType == '주장'): ws1.append(temp)
            elif(qType == '요지'):ws2.append(temp)
            elif (qType == '주제'):ws3.append(temp)
            elif (qType == '제목'):ws4.append(temp)
            elif (qType == '심경_분위기_변화'):ws5.append(temp)
            elif (qType == '빈칸(구,절)'):ws6.append(temp)
            elif(qType == '빈칸(단어)'): ws7.append(temp)

            # 빈칸(AB) ?

            elif (qType == '연결사'):ws8.append(temp)
            elif(qType == '요약(구,절)'):ws9.append(temp)
            elif(qType == '요약(단어)'): ws10.append(temp)


wb.save(filename='C:/Users/LXPER MINI001/Desktop/잡업/{}.xlsx'.format(fName))
