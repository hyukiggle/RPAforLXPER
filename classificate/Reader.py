# EBS 문제 엑셀 유형별 분류 코드

from openpyxl import Workbook
import os
import glob

# 파일이 달라짐에 따라 경로나 이름을 설정해주어야함.

os.chdir("//lxper-share/LXPER공유폴더/이승혁/training data/E/주제/txt")
filelist = os.listdir()
for files in filelist:
    if files.endswith(".txt"):

        file = open("//lxper-share/LXPER공유폴더/이승혁/training data/E/주제/txt/{}".format(files),'r', encoding='UTF8')
        totalLines = file.readlines()

        valid_index = []
        allLines = []
        questionLines = []
        n = 1
        pointLine = []
        pointIndex = []
        starLine = []
        starIndex = []
        refIndex = []
        sharpIndex =[]
        arrowIndex = []
        subject = []
        for line in totalLines:
            allLines.append(line.strip())
            # 파일이 달라짐에 따라 경로나 이름을 설정해주어야함.
            if(('네모 안에서'in line) or ('#' in line) or ('밑줄 친' in line) or ('흐름으로 보아' in line)):
                if(('분류' not in line) and ('>' not in line) and ('출처' not in line) and ('어법' not in line)):
                    questionLines.append(line)

                    print(line)
                    valid_index.append(n)
                # if ('#' in line):
                #     sharpIndex.append(n)

            if('점]' in line):
                pointLine.append(line)
                pointIndex.append(n)
            elif('*' in line):
                starLine.append(line)
                starIndex.append(n)
            elif('↓' in line):
                arrowIndex.append(n)
            if('출처' in line):
                refIndex.append(line[4:].strip())

            n+=1
        '''
        for q in range(len(questionLines)):
            questionLine = questionLines[q]
            for u in range(len(questionLines[q])):
                if(questionLine[u] == '.'):
                    questionLine = questionLine[u:]
                    ws.cell(row=q+2,column=2).value = questionLine.strip()
                    ws.cell(row=q+2,column=1).value = questionLine[:u]
        '''
        print(len(questionLines),questionLines)
        print(starIndex)
        print(len(valid_index), len(pointIndex))
        print(pointIndex)
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value='번호'
        ws.cell(row=1, column=2).value='문제'
        ws.cell(row=1, column=3).value='지문'
        ws.cell(row=1,column=4).value = '정답'
        ws.cell(row=1, column=5).value='보기1'
        ws.cell(row=1, column=6).value='보기2'
        ws.cell(row=1, column=7).value='보기3'
        ws.cell(row=1, column=8).value='보기4'
        ws.cell(row=1, column=9).value='보기5'
        ws.cell(row=1,column=10).value = '별단어'
        ws.cell(row=1, column=11).value = '출처'


        #문제
        for q in range(2, len(questionLines)+2):
            for u in range(len(questionLines[q-2])):
                if(questionLines[q-2][u] == '.'):
                    ws.cell(row=q, column=2).value = questionLines[q - 2][u+1:].strip()
                    ws.cell(row=q, column=1).value = questionLines[q - 2][:u+1]
                if(questionLines[q-2][u] == "'"):
                    for s in range(u,len(questionLines[q-2])):
                        if (questionLines[q - 2][s] == "의"):
                            subject.append(questionLines[q-2][u:s])
        for sub in range(len(subject)):
            ws.cell(row=q+2,column=13).value = subject[sub]

        #지문
        for i in range(len(valid_index)-1):
            #print(allLines[valid_index[i - 2]]) #:pointIndex[i - 1]])
            sentence = ""
            for j in range(valid_index[i],valid_index[i+1]):
                sentence += allLines[j].strip()
                #print(len(sentence), type(sentence))
                for s in range(len(sentence)):
                    if((sentence[s:s+2] == ".*") or (sentence[s:s+2] == '"*')):
                        almaeng = sentence[:s]
                        ws.cell(row=i + 2, column=3).value = almaeng
                        for e in range(s,len(sentence)):
                            if(sentence[e:e+1]=='['):
                                ws.cell(row=i + 2, column=10).value = sentence[s+1:e]
                    elif((sentence[s:s+1] == '[') or (sentence[s:s+2] =='.[')):
                        almaeng = sentence[:s]
                        ws.cell(row=i + 2, column=3).value = almaeng

                        #print(ast)

                        '''
                        ws.cell(row=i+2, column=3).value = sentence[:ast]
                    else:
                        ws.cell(row=i+2, column=3).value = sentence
                        '''
            print('-----------------------------')

        ws.cell(row=len(valid_index)+1, column=3).value = allLines[valid_index[-1]]
        ws.cell(row=len(valid_index)+1, column=10).value=allLines[valid_index[-1]+1]
        # 요약문
        for ar in range(len(arrowIndex)):
            summary = ""
            for row in range(arrowIndex[ar], pointIndex[ar]):
                summary += allLines[row]
            ws.cell(row=ar+2,column=4).value = summary

        # 선택지
        for j in range(2,len(pointIndex)+2):
            # print(allLines[pointIndex[j-2]+1:pointIndex[j-2]+11:2])
            answers = allLines[pointIndex[j-2]+1:pointIndex[j-2]+11:2]
            for ans in range(5):
                ws.cell(row=j, column=5).value = answers[0]
                ws.cell(row=j, column=6).value = answers[1]
                ws.cell(row=j, column=7).value = answers[2]
                ws.cell(row=j, column=8).value = answers[3]
                ws.cell(row=j, column=9).value = answers[4]

            # print(answers)

        #출처
        for c in range(2, len(refIndex)+2):
            ref = refIndex[c-2]
            ws.cell(row=c, column=11).value = ref
        #점수
        for p in range(len(pointIndex)):
            ws.cell(row=p+2, column=12).value = pointLine[p]
        print(sharpIndex)
        print(valid_index)
        # 파일이 달라짐에 따라 경로나 이름을 설정해주어야함.
        wb.save(filename = 'C:/Users/LXPER MINI001/Desktop/지문작업분류/{}.xlsx'.format(files[:-4]))