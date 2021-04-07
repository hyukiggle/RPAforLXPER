from openpyxl import load_workbook
import os

os.chdir("//lxper-share/LXPER공유폴더/이승혁/training data")
file = open("//lxper-share/LXPER공유폴더/이승혁/training data/TextMerged.txt",'r',encoding='UTF8')
wb = load_workbook("//lxper-share/LXPER공유폴더/이승혁/training data/EbsMergedInaRow.xlsx")
ws = wb.active


number_five = []
direction_line = []
index = 0
index_list = []
directionIndex = []
fiveIndex = []
contents = []
typeList = ['분위기','심경','심정','요지','주제','제목','주장','빈칸','요약','시사','쟁점','교훈','어조','어구', '목적','흐름','생각','태도','관점','속담','의도','의미','내용','의견','어법','문맥','구성','직업']

total = file.readlines()
for line in total:
    if ('-' and '.') in line:
        for telem in typeList:
            if telem in line:
                if '다음' in line or'#' in line or '위' in line or  '-' in line or   '윗' in line or '@' in line or '빈칸' in line or '가장' in line:
                    direction_line.append(line)
                    index_list.append(index)

#             elif '#' in line:
#                 direction_line.append(line)
#                 index_list.append(index)

    elif('⑤' in line):
        fiveIndex.append(index)

    index += 1

newIndex= []
newDir = []
newFive = []
# 중복제거 과정
for elem in index_list:
    if elem not in newIndex:
        newIndex.append(elem)
for d in direction_line:
    if d not in newDir:
        newDir.append(d)
for f in fiveIndex:
    if f not in newFive:
        newFive.append(f)

# 지문의 1,2줄만 가져오기
for i in range(len(newIndex)):
    contents.append(total[newIndex[i]+1].strip())

allC = []
# 모든 선택지
for d in range(len(newFive)):
    allC.append(total[newFive[d]- 7].strip())
    allC.append(total[newFive[d] - 5].strip())
    allC.append(total[newFive[d] - 3].strip())
    allC.append(total[newFive[d] - 1].strip())
    allC.append(total[newFive[d] + 1].strip())
len(allC)

for d in newDir:
    print(d)
print(len(newDir),len(newIndex),len(contents),len(newFive),len(allC))
'''
if(len(newIndex) > len(newFive)):
    for r in range(1,ws.max_row):
        # 지문
        print(r," start")
        for c in range(len(contents)):
            a = c + 2
            if(contents[c] in ws['A'][r].value):
                if(ws['E'][r].value == allC[a] and ws['F'][r].value == allC[a+1] and ws['G'][r].value == allC[a+2]):
                    print(r," ok")
                else:
                    print(ws['A'][r].value)
        print(r, " done")
else:
    print("failed")

'''
