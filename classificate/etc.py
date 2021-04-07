from openpyxl import Workbook
import os
wb = Workbook()
ws = wb.active
# 문제 유형분류(주제, 제목, 주장, 요지, 빈칸, 문단요약)
def qClassify(filename):
    qType = ""
    if('주제' in filename):
        qType = '주제'
    elif ('제목' in filename):
        qType = '제목'
    elif ('주장' in filename):
        qType = '주장'
    elif ('요지' in filename):
        qType = '요지'
    elif ('빈칸에 들어갈 ' in filename):
        qType = '빈칸'
    elif ('요약' in filename):
        qType = '문단요약'
    return qType
# 시트 행구분
def initialize(qType):
    if(qType == '문단요약'):
        ws.cell(row=1, column=1).value = '번호'
        ws.cell(row=1, column=2).value = '문제'
        ws.cell(row=1, column=3).value = '지문'
        ws.cell(row=1, column=4).value = '요약문'
        ws.cell(row=1, column=5).value = '정답'
        ws.cell(row=1, column=6).value = '보기1'
        ws.cell(row=1, column=7).value = '보기2'
        ws.cell(row=1, column=8).value = '보기3'
        ws.cell(row=1, column=9).value = '보기4'
        ws.cell(row=1, column=10).value = '보기5'
        ws.cell(row=1, column=11).value = '별단어'
        ws.cell(row=1, column=12).value = '출처'
    else:
        ws.cell(row=1, column=1).value = '번호'
        ws.cell(row=1, column=2).value = '문제'
        ws.cell(row=1, column=3).value = '지문'
        ws.cell(row=1, column=4).value = '정답'
        ws.cell(row=1, column=5).value = '보기1'
        ws.cell(row=1, column=6).value = '보기2'
        ws.cell(row=1, column=7).value = '보기3'
        ws.cell(row=1, column=8).value = '보기4'
        ws.cell(row=1, column=9).value = '보기5'
        ws.cell(row=1, column=10).value = '별단어'
        ws.cell(row=1, column=11).value = '출처'
# 요약문 모음 리스트
def summaryParser(totalLine):
    n = 0
    totalIndex = []
    arrowIndex = []
    summary = []
    for line in totalLine:
        totalIndex.append(n)
        if('↓' in line):
            arrowIndex.append(n)
        n+=1

    for i in range(len(arrowIndex)):
        summary.append(totalLine[arrowIndex[i]])
    return summary
# 선택지 모음 list
def optionCollector(totalLine,start, end):
    options = []
    for opt in range(start, end):
        if('①' in totalLine[opt]):
            options.append(totalLine[opt][2:])
            options.append(totalLine[opt+1][2:])
            options.append(totalLine[opt + 2][2:])
            options.append(totalLine[opt + 3][2:])
            options.append(totalLine[opt + 4][2:])
    return options
# 2행부터 끝까지 채워주는 function
def stuff(list,col):
    for i in range(len(list)):
        ws.cell(row= i+2, column=col).value = list[i]
# 선택지 모음 리스트를 채워주는 function
def stuffOption(list, col):
    for i in range(len(list)):
        rw = i // 5
        cl = i % 5
        if(cl == 0):
            ws.cell(row=rw+2,column=cl).value = list[i]
        elif(cl == 1):
            ws.cell(row=rw + 2, column=cl+1).value = list[i]
        elif (cl == 2):
            ws.cell(row=rw + 2, column=cl+2).value = list[i]
        elif (cl == 3):
            ws.cell(row=rw + 2, column=cl+3).value = list[i]
        elif (cl == 4):
            ws.cell(row=rw + 2, column=cl+4).value = list[i]
#별단어모으기
def starwordParser(sentence):
    starwords = []
    for i in range(len(sentence)):
        if(sentence[i:i+1] == '*'):
            starwords.append(sentence[i:])
        else:
            starwords.append('')
def printInfo(list):
    print(len(list),list)
os.chdir("//lxper-share/LXPER공유폴더/이승혁")
files = os.listdir()
for file in files:
    if file.endswith('.txt'):
        f = open(file, 'r', encoding='utf-8')
        print(file,"start")
        mType = qClassify(file)
        initialize(mType)

        totalLine = f.readlines()
        refs = []
        refNums = []
        answers = []
        directions = []
        dirIndex = []
        contents = []
        options = []
        starwords = []
        for i in range(len(totalLine)):
            if (('_고3' in totalLine[i])):  # total(i) 는 출처가 있는 라인
                refs.append(totalLine[i][:-4])
                refNums.append(totalLine[i].strip()[-2:])
                num = totalLine[i + 1].strip()[-1]
                if (num == '①'): answer = 1
                elif (num == '②'): answer = 2
                elif (num == '③'): answer = 3
                elif (num == '④'): answer = 4
                elif (num == '⑤'): answer = 5
                else: answer = ""
                answers.append(answer)
            if(mType in totalLine[i]):
                dirIndex.append(i)
                directions.append(totalLine[i].strip())

        for k in range(1, len(dirIndex)):
            for j in range(dirIndex[k-1], dirIndex[k]):

                if ((totalLine[j] == '') & (totalLine[j + 1] != '')):
                    sentence = ""
                    if ((') ①' not in totalLine[j + 1])
                            and ('① ' not in totalLine[j + 1])
                            and ('고' not in totalLine[j + 1])
                            and ('어' not in totalLine[j + 1])
                            and (') ④' not in totalLine[j + 1])
                            and (') ⑤' not in totalLine[j + 1])
                            and ('   (C)' not in totalLine[j + 1])):
                        sentence += totalLine[j + 1].strip()
                        print(sentence)
                        starwords.starwordParser(sentence)
                    contents.append(sentence)
                    options = optionCollector(totalLine,dirIndex[k-1], dirIndex[k])
            if(mType == '문단요약'):
                summary = summaryParser(totalLine)
        printInfo(refNums)
        printInfo(directions)
        printInfo(contents)
        printInfo(answers)
        printInfo(options)
        printInfo(refs)
        printInfo(starwords)
        if(mType == '문단요약'):
            stuff(refNums,1)
            stuff(directions,2)
            stuff(contents,3)
            stuff(summaryParser, 4)
            stuff(answers,5)
            stuffOption(options,6)
            stuff(starwords,11)
            stuff(refs, 12)
        else:
            stuff(refNums, 1)
            stuff(directions, 2)
            stuff(contents, 3)
            stuff(answers, 4)
            stuffOption(options, 5)
            stuff(starwords, 10)
            stuff(refs, 11)
        print(file, "end")
        wb.save('C:/Users/LXPER MINI001/Desktop/잡업/TEST/{}.xlsx'.format(file[:-4]))