# 평가원,교육청 기출문제 엑셀 유형별 분류 코드
import os
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 리스트의 요소들을 한 열에 쭉 채우는 method
def stuffit(list,col):
    for i in range(len(list)):
        ws.cell(row = i+2, column=col).value = list[i]

def printInfo(list):
    print(len(list), list)

# 만들고자하는 엑셀파일의 대상이 될 txt파일 경로 복사

os.chdir(r"//lxper-share/LXPER공유폴더/이승혁/기출/1학년")
filelist = os.listdir()
for files in filelist:
    if files.endswith("요지.txt"):
        print(files[:-4],"작업 시작")
        file = open("{}.txt".format(files[:-4]),'r', encoding='UTF8')
        ws.cell(row=1, column=1).value = '번호'
        ws.cell(row=1, column=2).value = '문제'
        ws.cell(row=1, column=3).value = '지문'
        ws.cell(row=1, column=4).value = '정답'
        ws.cell(row=1, column=5).value = '보기1'
        ws.cell(row=1, column=6).value = '보기2'
        ws.cell(row=1, column=7).value = '보기3'
        ws.cell(row=1, column=8).value = '보기4'
        ws.cell(row=1, column=9).value = '보기5'
        ws.cell(row=1, column=10).value = '별단어'
        ws.cell(row=1, column=12).value = '출처'
        ws.cell(row=1,column=11).value = '요약문'

        directions = []
        dirIndex = []
        contents = []
        starIndex=[]
        starLine = []
        answers = []
        refs = []
        refNums = []
        options = []
        total = file.readlines()
        lines = []
        additionalStar = []
        for line in total:
            lines.append(line.strip())
        for i in range(len(lines)-1):
            if(('_고1' in lines[i])):  # total(i) 는 출처가 있는 라인
                refs.append(lines[i][:-3])
                refNums.append(lines[i][-2:])

                num = lines[i+1][-1]
                if (num == '①'): answer = 1
                elif (num == '②'): answer = 2
                elif (num == '③'): answer = 3
                elif (num == '④'): answer = 4
                elif (num == '⑤'): answer = 5
                else: answer = ""

                answers.append(answer)
                '''
                if('주제' in total[i+1]):
                    type = '주제'
                elif ('제목' in total[i + 1]):
                    type = '주제'
                elif ('요지' in total[i + 1]):
                    type = '주제'
                elif ('주장' in total[i + 1]):
                    type = '주제'
                elif ('요약' in total[i + 1]):
                    type = '주제'
                elif('빈칸' in total[i + 1]):
                    if(total[i+1][-2]=='@'):
                        type = '빈칸(단어)'
                    elif(total[i+1][-2] == 'T'):
                        type = '연결어'
                    else:
                        type = '빈칸(구,절)'
                '''
                directions.append(lines[i+1])
                refs.append(lines[i-1])
                dirIndex.append(i)




        newDir = []
        newAns = []
        newCont = []

        for i in range(1,len(dirIndex)):
            for j in range(dirIndex[i-1], dirIndex[i]):
                if((lines[j] == '') & (lines[j+1] != '')):
                    sentence = ""
                    if((') ①' not in lines[j+1])
                            and ('① ' not in lines[j+1])
                            and ('고' not in lines[j+1])
                            and ('어' not in lines[j+1])
                            and(') ④' not in lines[j+1])
                            and(') ⑤' not in lines[j+1])
                            and('   (C)'not in lines[j+1])):

                        sentence += lines[j+1].strip()
                    contents.append(sentence)
                    if ('① ' in lines[j + 1]):   #선택지 모음
                        options.append(lines[j + 1][2:])
                        options.append(lines[j + 2][2:])
                        options.append(lines[j + 3][2:])
                        options.append(lines[j + 4][2:])
                        options.append(lines[j + 5][2:])
                    for star in range(len(sentence)):
                        if(sentence[star:star+1] == '*'):
                            starword = sentence[star:]
                            starLine.append(starword)
                            ws.cell(row=i+1, column=10).value = starword
                    if('*' in line[:4]):
                        additionalStar.append(line)
                    elif('*' not in line[:4]):
                        additionalStar.append(" ")


        onec = []
        twoc = []
        threec = []
        fourc = []
        fivec = []
        for a in range(len(options)):
            rem = a % 5
            if (rem == 0):
                onec.append(options[a])
            elif (rem == 1):
                twoc.append(options[a])
            elif (rem == 2):
                threec.append(options[a])
            elif (rem == 3):
                fourc.append(options[a])
            elif (rem == 4):
                fivec.append(options[a])


        for c in contents:
            if c == "":
                contents.remove(c)
            if c =='\n':
                contents.remove(c)
        for r in refs:
            if r == '':
                refs.remove(r)
        '''
        for k in range(len(newDir)):
            ws.cell(row=k+2, column=2).value = newDir[k]
            if('어법상 틀린' in newDir[k]):  #어법
                ws.cell(row=k+2, column=1).value = '어법'
            elif(('문맥상 낱말의' in newDir[k]) or ('낱말로' in newDir[k])):   #어휘
                ws.cell(row=k + 2, column=1).value = '어휘'
            elif('빈칸에 들어갈 말로' in newDir[k]):    #빈칸
                ws.cell(row=k + 2, column=1).value = '빈칸'
        '''


        printInfo(refNums)
        printInfo(directions)
        printInfo(contents)
        printInfo(answers)
        printInfo(options)
        printInfo(refs)

        stuffit(refNums,1)
        stuffit(directions,2)
        stuffit(contents,3)
        stuffit(answers,4)
        stuffit(onec,5)
        stuffit(twoc,6)
        stuffit(threec,7)
        stuffit(fourc,8)
        stuffit(fivec,9)
        #stuffit(starLine,10)
        stuffit(refs,12)
        #stuffit(additionalStar,10)

        print(files[:-4],"작업 끝")
        del contents
        del directions
        del refNums
        del answers

        wb.save(filename="C:/Users/LXPER MINI001/Desktop/잡업/{}.xlsx".format(files[:-4]))

