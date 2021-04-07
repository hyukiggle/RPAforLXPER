# EBS 문제 엑셀 유형별 분류 코드

from openpyxl import Workbook



# 파일이 달라짐에 따라 경로나 이름을 설정해주어야함.

file = open("//lxper-share/LXPER공유폴더/이승혁/기출/2학년/2010_2020년_고2_교육청 모의고사_요약문.txt",'r',encoding='utf=8')
totalLines = file.readlines()

valid_index = []
allLines = []
questionLines = []
n = 1
starLine = []
starIndex = []
refs = []
refNum =[]
arrowIndex = []
answers = []
for line in totalLines:
    allLines.append(line.strip())
    # 파일이 달라짐에 따라 경로나 이름을 설정해주어야함.
    if('_고1' in line):
        ref = line[:-4]
        refs.append(ref)
        num = line[-3:]
        refNum.append(num)
    if('요약하고자' in line):
        questionLines.append(line)
        num = line.strip()[-1]
        if (num == '①'): answer = 1
        elif (num == '②'): answer = 2
        elif (num == '③'): answer = 3
        elif (num == '④'): answer = 4
        elif (num == '⑤'): answer = 5
        else: answer = ""
        answers.append(answer)
        print(line)
        valid_index.append(n)
    elif ('↓' in line):
        arrowIndex.append(n)
    elif('*' in line):
        starLine.append(line)
        starIndex.append(n)

    n+=1

print(len(questionLines),questionLines)
print(starIndex)
wb = Workbook()
ws = wb.active
ws.cell(row=1, column=1).value='번호'
ws.cell(row=1, column=2).value='문제'
ws.cell(row=1, column=3).value='지문'
ws.cell(row=1,column=4).value = '요약문'
ws.cell(row=1, column=5).value='정답'
ws.cell(row=1, column=6).value='보기1'
ws.cell(row=1, column=7).value='보기2'
ws.cell(row=1, column=8).value='보기3'
ws.cell(row=1, column=9).value='보기4'
ws.cell(row=1,column=10).value = '보기5'
ws.cell(row=1, column=11).value = '별단어'
ws.cell(row=1,column=12).value = '출처'

#문제
for q in range(2, len(questionLines)+2):
    ws.cell(row=q, column=2).value = questionLines[q - 2].strip()

contents = []
options  = []
summary = []
#지문
for i in range(1, len(valid_index)):
    for j in range(valid_index[i - 1], arrowIndex[i-1]):
        sentence = ""
        if ((allLines[j] == '') & (allLines[j + 1] != '')):
            sentence += allLines[j + 1]

        for star in range(len(sentence)):
            if(sentence[star:star+1] == '*'):
                ws.cell(row=i+2, column=11).value = sentence[star:]
                ws.cell(row=i + 2, column=3).value = sentence[:star]
            else:
                ws.cell(row=i + 2, column=3).value = sentence
        contents.append(sentence)

    summary.append(allLines[arrowIndex[i-1]])

    for k in range(arrowIndex[i-1],valid_index[i]):
        if ('① ' in allLines[k]):  # 선택지 모음
            options.append(allLines[k][2:])
            options.append(allLines[k+1][2:])
            options.append(allLines[k+2][2:])
            options.append(allLines[k+3][2:])
            options.append(allLines[k+4][2:])
    onec = []
    twoc = []
    threec = []
    fourc = []
    fivec = []
    for a in range(len(options)):
        rem = a % 5
        if (rem == 0):
            onec.append(options[a])
        elif (rem == 1):
            twoc.append(options[a])
        elif (rem == 2):
            threec.append(options[a])
        elif (rem == 3):
            fourc.append(options[a])
        elif (rem == 4):
            fivec.append(options[a])



    #ws.cell(row=i+2, column=12).value = refs[i]
    #ws.cell(row=i+2,column=1).value = refNum[i]

    print('-----------------------------')
for con in contents:
    if con == '\n':
        contents.remove(con)
    print(con.strip())
print(answers)
for i in range(len(fivec)):
    ws.cell(row=i + 2, column=6).value = onec[i]
    ws.cell(row=i + 2, column=7).value = twoc[i]
    ws.cell(row=i + 2, column=8).value = threec[i]
    ws.cell(row=i + 2, column=9).value = fourc[i]
    ws.cell(row=i + 2, column=10).value = fivec[i]
    ws.cell(row=i+2, column=4).value=summary[i]
    #ws.cell(row=i+2, column=3).value = contents[i]
    ws.cell(row=i + 2, column=5).value = answers[i]
ws.cell(row=len(valid_index)+1, column=3).value = allLines[valid_index[-1]]
ws.cell(row=len(valid_index)+1, column=10).value=allLines[valid_index[-1]+1]
# 요약문
#for ar in range(len(arrowIndex)):



# 파일이 달라짐에 따라 경로나 이름을 설정해주어야함.
wb.save(filename = 'C:/Users/LXPER MINI001/Desktop/잡업/2010_2020년_고2_교육청 모의고사_요약문temp.xlsx')


