from bs4 import BeautifulSoup
import requests
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
import os

wb = Workbook()
ws = wb.active
'''
options = wb.ChromeOptions()
options.add_argument('headless')
options.add_argument('window-size=1920x1080')
options.add_argument("disable-gpu")
'''
driver = webdriver.Chrome('chromedriver.exe')
driver.execute_script('window.open("about:blank", "_blank");')

tabs = driver.window_handles


# TAB_1
driver.switch_to_window(tabs[0])
driver.get("http://www.ebsi.co.kr/ebs/pot/potl/login.ebs?destination=/index.jsp&alertYn=N")
driver.find_element_by_id('userid').send_keys('patrickjane4')
driver.find_element_by_id('textfield3').send_keys('dudahf43001')
driver.find_element_by_xpath('//*[@id="reNcontents"]/form/div[2]/div/fieldset/div[1]/button').click()

#TAB_2
mainURL = "http://www.ebsi.co.kr/ebs/ai/xipa/MyPaper.ebs"
driver.switch_to_window(tabs[0])
driver.get(mainURL)
#driver.find_element_by_xpath('//*[@id="paperFrm"]/div[1]/ul/li[1]/span[2]/div[2]/button[1]').click()
#driver.implicitly_wait(3)
#parent_window = driver.current_window_handle
#all_windows = driver.window_handles
#child_window = [window for window in all_windows if window != parent_window][0]

total_window = driver.window_handles

#

#print(url)
time.sleep(5)
questionNum = []
for i in range(44):
    questionNum.append(i)
for p in range(2,6):
        pageNext = driver.find_element_by_class_name("pageNext")
        button = pageNext.find_element_by_css_selector('a')
        button.click()

qList = []
url = driver.current_url
pageSource = requests.get(mainURL).text
list02 = driver.find_element_by_css_selector('#paperFrm > div.resultDiv > ul')
testList = list02.find_elements_by_tag_name(('li'))
allAnswers =[]
for go in range(1,len(testList)):


    testName = testList[go].find_element_by_class_name('q_tit').text
    qList.append(testName)
    print(testName)
    # 문제지 열기
    testList[go].find_element_by_xpath('//*[@id="paperFrm"]/div[1]/ul/li[{}]/span[2]/div[2]/button[2]'.format(go+1)).send_keys(Keys.ENTER)
    driver.switch_to_window(driver.window_handles[-1])
    url = driver.current_url
    time.sleep(5)
    #print(url)
    driver.find_element_by_xpath('//*[@id="wrap"]/div[2]/div/div[2]/div[4]/span/a').send_keys(Keys.ENTER)
    driver.switch_to_alert().accept()

    ansSheet = driver.find_element_by_id('omrbox')

    questions = ansSheet.find_elements_by_tag_name('tr')    # 전체 문제
    print(len(questions))
    answers = []
    for index in range(0,len(questions)):
        omrNumber = 'omr_answer_{}'.format(index+1)
        sheet = questions[index].find_elements_by_tag_name('td')[1]
        onetofive = sheet.find_elements_by_class_name(omrNumber)
        #print(omrNumber,len(onetofive))
        for n in range(len(onetofive)):
            #temp = onetofive[n].get_attribute('iscorrectanswer')
            if(onetofive[n].get_attribute('iscorrectanswer') == 'true'):
                answers.append(n+1)
    print(len(answers),answers)
    allAnswers.append(answers)

    driver.switch_to_window(driver.window_handles[-2])
    driver.switch_to_window(tabs[0])


    #print(questions[0].find_elements_by_tag_name('td')[1].find_element_by_class_name('omr_answer_1').get_attribute('iscorrectanswer'))
print(len(qList),len(allAnswers))
for i in range(len(allAnswers)):
    print(len(allAnswers[i]))
os.chdir("C:/Users/LXPER MINI001/Desktop/잡업")
for filename in range(len(qList)):
    ws.cell(row=1, column = 1+filename).value = qList[filename]
for big in range(len(allAnswers)):
    eachTest = allAnswers[big]
    for ans in range(len(eachTest)):
        ws.cell(row=ans+2,column=1+big).value = eachTest[ans]
wb.save(filename="EBS빈칸정답모음.xlsx")